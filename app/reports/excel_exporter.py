import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_codes_excel(codes: list, batch_name: str, subject_name: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Access Codes"

    # Styles
    hdr_fill = PatternFill("solid", fgColor="0F2D6B")
    alt_fill = PatternFill("solid", fgColor="EFF6FF")
    hdr_font = Font(name="Cairo", bold=True, color="FFFFFF", size=11)
    body_font= Font(name="Cairo", size=10)
    center   = Alignment(horizontal="center", vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{subject_name} — {batch_name}"
    ws["A1"].font      = Font(name="Cairo", bold=True, size=14, color="0F2D6B")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font      = Font(name="Cairo", size=9, color="888888")
    ws["A2"].alignment = center

    # Headers
    headers   = ["#", "Code", "Subject", "Batch", "Max Uses", "Used", "Expiry", "Status"]
    col_widths = [4, 16, 22, 18, 10, 8, 16, 10]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=col)
        c.value     = h
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 28

    # Data rows
    for i, code in enumerate(codes, 1):
        row_num = i + 3
        shade   = alt_fill if i % 2 == 0 else None
        status  = "Active" if code.get("is_active", True) else "Inactive"
        expiry  = str(code.get("expires_at",""))[:10] if code.get("expires_at") else "No Expiry"
        row_data = [
            i,
            code.get("code",""),
            subject_name,
            batch_name,
            code.get("max_uses", 1) or "∞",
            code.get("uses_count", 0),
            expiry,
            status,
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.font      = Font(name="Cairo", bold=(col==2), size=10,
                               color="16A34A" if (col==8 and status=="Active")
                               else "DC2626" if col==8 else "111827")
            c.alignment = center
            c.border    = border
            if shade:
                c.fill = shade
        ws.row_dimensions[row_num].height = 22

    # Summary
    sr = len(codes) + 5
    ws.cell(row=sr, column=1).value = f"Total: {len(codes)}"
    ws.cell(row=sr, column=1).font  = Font(name="Cairo", bold=True, color="0F2D6B")
    used = sum(1 for c in codes if c.get("uses_count",0) >= c.get("max_uses",1) > 0)
    ws.cell(row=sr, column=2).value = f"Used: {used}"
    ws.cell(row=sr, column=3).value = f"Available: {len(codes)-used}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{subject_name}_{batch_name}.xlsx".replace(" ", "_")
    return buf, fname


def export_results_excel(results: list, title: str = "EduVision Results"):
    """Flat results export — reused from v2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    hdr_fill = PatternFill("solid", fgColor="0F2D6B")
    hdr_font = Font(name="Cairo", bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EFF6FF")
    center   = Alignment(horizontal="center", vertical="center")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers   = ["#","Student","Exam","Subject","Score","Total","%","Grade","Date"]
    col_widths = [4, 22, 25, 18, 8, 8, 8, 8, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col)
        c.value     = h
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, r in enumerate(results, 1):
        pct   = round(r.get("score",0)/r.get("total",1)*100) if r.get("total") else 0
        grade = "A" if pct>=90 else "B" if pct>=80 else "C" if pct>=70 else "D" if pct>=60 else "F"
        shade = alt_fill if i%2==0 else None
        row   = [i, r.get("student_name","—"), r.get("exam_title","—"),
                 r.get("subject_name","—"), r.get("score",0), r.get("total",0),
                 f"{pct}%", grade, str(r.get("submitted_at",""))[:10]]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i+1, column=col, value=val)
            c.alignment = center
            c.border    = border
            c.font      = Font(name="Cairo", size=10)
            if shade: c.fill = shade

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
