import streamlit as st
import re
import pandas as pd
from app.core.security import get_supabase, get_supabase_admin, get_current_user_id
from translations import get_lang


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def safe_color(hex_str: str) -> str:
    """XSS-safe hex color — strips anything that isn't 0-9 A-F."""
    clean = re.sub(r'[^0-9A-Fa-f]', '', str(hex_str or ""))
    return clean[:6] if len(clean) >= 6 else "3B82F6"


def sanitize(val, max_len: int = 200) -> str:
    """Strip HTML-dangerous characters from user input."""
    return re.sub(r'[<>&"\']', '', str(val or "").strip())[:max_len]


PAGE_SIZE = 50   # students per page


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────
def show_admin():
    lang = get_lang()
    sb   = get_supabase()
    uid  = get_current_user_id()

    with st.sidebar:
        st.markdown(f"### 👑 {'لوحة التحكم' if lang=='ar' else 'Admin Panel'}")

        groups = {
            "الرئيسية": [
                ("dashboard",    "📊", "الرئيسية"),
                ("results",      "📈", "النتائج"),
                ("audit",        "📋", "سجل الأحداث"),
            ],
            "الإدارة": [
                ("users",        "👥", "المستخدمين"),
                ("subjects",     "📚", "المواد"),
                ("access_codes", "🔑", "أكواد الوصول"),
            ],
            "العمليات": [
                ("attendance",   "📋", "الحضور"),
                ("sales",        "💰", "المبيعات"),
                ("security",     "🔒", "الأمان"),
            ],
        }

        if "admin_page" not in st.session_state:
            st.session_state.admin_page = "dashboard"

        for group, items in groups.items():
            st.caption(group)
            for key, icon, label in items:
                active = st.session_state.admin_page == key
                if st.button(f"{icon} {label}", use_container_width=True,
                             type="primary" if active else "secondary",
                             key=f"admin_nav_{key}"):
                    st.session_state.admin_page = key
                    st.session_state.pop("student_page", None)
                    st.rerun()

    page = st.session_state.admin_page

    if   page == "dashboard":    _dashboard(sb, lang)
    elif page == "users":        _users(sb, lang, uid)
    elif page == "subjects":     _subjects(sb, lang, uid)
    elif page == "access_codes": _access_codes(sb, lang, uid)
    elif page == "attendance":
        from app.pages.admin.attendance import show_attendance
        show_attendance()
    elif page == "sales":
        from app.pages.admin.sales import show_sales
        show_sales()
    elif page == "security":
        from app.pages.admin.security import show_security
        show_security()
    elif page == "results":      _results(sb, lang)
    elif page == "audit":        _audit(sb, lang)


# ─────────────────────────────────────────────────────────
# DASHBOARD  (PERF-02: single RPC call instead of 6+)
# ─────────────────────────────────────────────────────────
def _dashboard(sb, lang):
    st.title("📊 الرئيسية")

    # One RPC call replaces 6 sequential queries
    try:
        stats = get_supabase_admin().rpc("get_admin_dashboard_stats").execute().data or {}
    except Exception:
        stats = {}

    students    = stats.get("students", 0)
    teachers    = stats.get("teachers", 0)
    subjects    = stats.get("subjects", 0)
    enrollments = stats.get("enrollments", 0)
    revenue     = stats.get("revenue", 0)
    alerts      = stats.get("alerts", 0)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("🎓 الطلاب",     students)
    c2.metric("📚 المعلمون",   teachers)
    c3.metric("📖 المواد",     subjects)
    c4.metric("📝 التسجيلات", enrollments)
    c5.metric("💰 الإيرادات", f"{revenue} ج.م")

    if alerts > 0:
        st.warning(f"⚠️ {alerts} تنبيه أمني يحتاج مراجعة")

    st.divider()
    st.subheader("📈 آخر النتائج")
    try:
        res = sb.table("results") \
                .select("submitted_at, score, total, profiles(full_name), exams(title)") \
                .order("submitted_at", desc=True).limit(10).execute()
        if res.data:
            rows = []
            for r in res.data:
                pct = round(r["score"] / r["total"] * 100) if r["total"] else 0
                rows.append({
                    "الطالب":   (r.get("profiles") or {}).get("full_name", "—"),
                    "الامتحان": (r.get("exams") or {}).get("title", "—"),
                    "الدرجة":   f"{r['score']}/{r['total']} ({pct}%)",
                    "التاريخ":  str(r["submitted_at"])[:10],
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.info("لا توجد نتائج بعد")
    except Exception as e:
        st.error(f"خطأ: {e}")


# ─────────────────────────────────────────────────────────
# USERS  (PERF-03: server-side search + pagination)
# ─────────────────────────────────────────────────────────
def _users(sb, lang, uid):
    st.title("👥 المستخدمين")

    tab_staff, tab_students = st.tabs([
        "👨‍💼 الموظفون (معلمين / أدمن / كاشير)",
        "🎓 الطلاب",
    ])

    # ── Staff ────────────────────────────────────────────
    with tab_staff:
        with st.expander("➕ إضافة موظف جديد"):
            with st.form("add_user"):
                c1, c2 = st.columns(2)
                with c1:
                    u_name  = st.text_input("الاسم الكامل")
                    u_email = st.text_input("البريد الإلكتروني")
                with c2:
                    u_role = st.selectbox("الدور", ["teacher", "co_admin", "admin", "cashier"])
                    u_pass = st.text_input("كلمة المرور", type="password")
                if st.form_submit_button("➕ إضافة", use_container_width=True):
                    if all([u_name, u_email, u_pass]):
                        try:
                            from app.services.auth_service import AuthService
                            AuthService().admin_create_user(
                                sanitize(u_email), u_pass,
                                sanitize(u_name), u_role, uid
                            )
                            st.success("✅ تم إضافة المستخدم")
                            st.rerun()
                        except Exception as e:
                            st.error(str(e))
                    else:
                        st.error("أدخل جميع البيانات")

        try:
            users     = sb.table("profiles").select("id, full_name, account_status, created_at").execute().data or []
            roles_res = sb.table("user_roles").select("user_id, roles(name)").execute().data or []
            role_map  = {r["user_id"]: r["roles"]["name"] for r in roles_res if r.get("roles")}
            staff     = [u for u in users if role_map.get(u["id"], "student") != "student"]
            if staff:
                df = pd.DataFrame([{
                    "الاسم":           u["full_name"],
                    "الدور":           role_map.get(u["id"], "—"),
                    "الحالة":          u.get("account_status", "active"),
                    "تاريخ الانضمام":  str(u["created_at"])[:10],
                } for u in staff])
                st.dataframe(df, use_container_width=True, hide_index=True)
            else:
                st.info("لا يوجد موظفون مسجلون.")
        except Exception as e:
            st.error(f"خطأ: {e}")

    # ── Students  ────────────────────────────────────────
    with tab_students:
        with st.expander("➕ إضافة طالب"):
            with st.form("enroll_student"):
                c1, c2 = st.columns(2)
                with c1:
                    s_name  = st.text_input("الاسم الكامل")
                    s_email = st.text_input("البريد الإلكتروني")
                with c2:
                    s_phone = st.text_input("رقم الهاتف (اختياري)")
                    s_pass  = st.text_input("كلمة المرور", type="password")
                try:
                    grades  = sb.table("grades").select("*").order("order_num").execute().data or []
                    g_opts  = {g["name_ar"]: g["id"] for g in grades}
                    s_grade = st.selectbox("الصف", list(g_opts.keys()) if g_opts else ["—"])
                except Exception:
                    g_opts  = {}
                    s_grade = st.text_input("الصف الدراسي")

                if st.form_submit_button("✅ تسجيل", use_container_width=True, type="primary"):
                    if all([s_name, s_email, s_pass]):
                        try:
                            from app.services.auth_service import AuthService
                            AuthService().admin_create_user(
                                sanitize(s_email), s_pass,
                                sanitize(s_name), "student", uid,
                                extra={"phone": sanitize(s_phone), "grade_id": g_opts.get(s_grade)}
                            )
                            st.success(f"✅ تم تسجيل: {sanitize(s_name)}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"خطأ: {e}")
                    else:
                        st.error("أدخل الاسم والإيميل وكلمة المرور.")

        with st.expander("📥 رفع Excel لتسجيل جماعي"):
            st.info("الملف يحتوي على أعمدة: **full_name, email, password, phone (اختياري)**")
            uploaded = st.file_uploader("اختر ملف Excel", type=["xlsx"])
            if uploaded:
                try:
                    import openpyxl, io
                    wb      = openpyxl.load_workbook(io.BytesIO(uploaded.read()))
                    ws      = wb.active
                    headers = [str(c.value or "").strip().lower()
                               for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    rows    = list(ws.iter_rows(min_row=2, values_only=True))
                    st.write(f"📊 {len(rows)} طالب في الملف")
                    if st.button("🚀 بدء التسجيل الجماعي", type="primary"):
                        from app.services.auth_service import AuthService
                        auth = AuthService()
                        success_count, failed_count = 0, 0
                        prog = st.progress(0)
                        for i, row in enumerate(rows):
                            data = dict(zip(headers, row))
                            raw_email    = str(data.get("email", "")).strip().lower()
                            raw_name     = sanitize(str(data.get("full_name", "")))
                            raw_password = str(data.get("password", "")).strip()
                            raw_phone    = sanitize(str(data.get("phone", "")))

                            if not re.match(r'^[^@]+@[^@]+\.[^@]+$', raw_email):
                                failed_count += 1
                                prog.progress((i + 1) / len(rows))
                                continue
                            try:
                                auth.admin_create_user(
                                    raw_email, raw_password,
                                    raw_name, "student", uid,
                                    extra={"phone": raw_phone}
                                )
                                success_count += 1
                            except Exception:
                                failed_count += 1
                            prog.progress((i + 1) / len(rows))
                        st.success(f"✅ تم تسجيل {success_count} طالب")
                        if failed_count:
                            st.warning(f"⚠️ فشل {failed_count} (إيميل مكرر أو بيانات ناقصة)")
                except Exception as e:
                    st.error(f"خطأ في قراءة الملف: {e}")

        st.divider()
        st.subheader("📋 قائمة الطلاب")

        # Server-side search + pagination (PERF-03)
        try:
            roles_res   = sb.table("user_roles").select("user_id, roles(name)").execute().data or []
            student_ids = [r["user_id"] for r in roles_res
                           if (r.get("roles") or {}).get("name") == "student"]

            if not student_ids:
                st.info("لا يوجد طلاب مسجلون بعد.")
                return

            search = st.text_input("🔍 بحث باسم الطالب")
            if "student_page" not in st.session_state:
                st.session_state.student_page = 0
            current_page = st.session_state.student_page
            offset = current_page * PAGE_SIZE

            # Server-side filter + paginate
            q = get_supabase_admin() \
                .table("profiles") \
                .select("id, full_name, account_status, created_at", count="exact") \
                .in_("id", student_ids) \
                .order("full_name")
            if search:
                q = q.ilike("full_name", f"%{search}%")
            result  = q.range(offset, offset + PAGE_SIZE - 1).execute()
            students = result.data or []
            total    = result.count or 0

            st.caption(f"إجمالي: {total} طالب — صفحة {current_page + 1} / {max(1, -(-total // PAGE_SIZE))}")

            # Pagination controls
            pc1, pc2, pc3 = st.columns([1, 3, 1])
            with pc1:
                if current_page > 0 and st.button("⬅️ السابق"):
                    st.session_state.student_page -= 1
                    st.rerun()
            with pc3:
                if offset + PAGE_SIZE < total and st.button("التالي ➡️"):
                    st.session_state.student_page += 1
                    st.rerun()

            # Student cards
            for s in students:
                with st.container(border=True):
                    c1, c2, c3 = st.columns([4, 2, 2])
                    with c1:
                        st.markdown(f"**👤 {s['full_name']}**")
                        st.caption(f"انضم: {str(s['created_at'])[:10]}")
                    with c2:
                        status = s.get("account_status", "active")
                        st.markdown("🟢 نشط" if status == "active" else "🔴 محظور")
                    with c3:
                        if status == "active":
                            if st.button("🚫 حظر", key=f"ban_{s['id']}"):
                                get_supabase_admin().table("profiles") \
                                    .update({"account_status": "banned"}) \
                                    .eq("id", s["id"]).execute()
                                st.rerun()
                        else:
                            if st.button("✅ تفعيل", key=f"unban_{s['id']}"):
                                get_supabase_admin().table("profiles") \
                                    .update({"account_status": "active"}) \
                                    .eq("id", s["id"]).execute()
                                st.rerun()

        except Exception as e:
            st.error(f"خطأ: {e}")


# ─────────────────────────────────────────────────────────
# SUBJECTS
# ─────────────────────────────────────────────────────────
def _subjects(sb, lang, uid):
    st.title("📚 المواد الدراسية")

    with st.expander("➕ إضافة مادة"):
        with st.form("add_subject"):
            c1, c2 = st.columns(2)
            with c1:
                s_ar    = st.text_input("اسم المادة (عربي)")
                s_en    = st.text_input("Subject Name (English)")
            with c2:
                grades  = sb.table("grades").select("*").order("order_num").execute().data or []
                g_opts  = {g["name_ar"]: g["id"] for g in grades}
                s_grade = st.selectbox("الصف", list(g_opts.keys()))
                s_color = st.color_picker("اللون", "#3B82F6")
            if st.form_submit_button("➕ إضافة", use_container_width=True):
                if s_ar:
                    sb.table("subjects").insert({
                        "name_ar":   sanitize(s_ar),
                        "name_en":   sanitize(s_en or s_ar),
                        "grade_id":  g_opts.get(s_grade),
                        "color_hex": safe_color(s_color.replace("#", "")),
                    }).execute()
                    st.success("✅ تمت الإضافة")
                    st.rerun()
                else:
                    st.error("أدخل اسم المادة")

    try:
        teacher_role_id  = sb.table("roles").select("id").eq("name", "teacher").single().execute().data["id"]
        teacher_user_ids = [r["user_id"] for r in
                            sb.table("user_roles").select("user_id").eq("role_id", teacher_role_id).execute().data or []]
        all_teachers     = sb.table("profiles").select("id, full_name").in_("id", teacher_user_ids).execute().data or [] \
                           if teacher_user_ids else []
        teacher_map      = {t["id"]: t["full_name"] for t in all_teachers}
        subs             = sb.table("subjects").select("*, grades(name_ar)").eq("is_active", True).execute().data or []

        for sub in subs:
            sub_id        = sub["id"]
            assigned      = sb.table("subject_teachers").select("teacher_id").eq("subject_id", sub_id).execute().data or []
            assigned_ids  = [a["teacher_id"] for a in assigned]
            assigned_names = [teacher_map.get(tid, "—") for tid in assigned_ids if tid in teacher_map]
            color_display  = safe_color(sub.get("color_hex", ""))

            with st.expander(f"**{sub['name_ar']}** · {(sub.get('grades') or {}).get('name_ar','—')} · 🎨 #{color_display}"):
                c1, c2 = st.columns([3, 1])
                with c1:
                    st.markdown(f"**المعلمون الحاليون:** {', '.join(assigned_names) if assigned_names else '—'}")
                    if all_teachers:
                        avail = {t["full_name"]: t["id"] for t in all_teachers}
                        sel   = st.multiselect("تعيين المعلمين", list(avail.keys()),
                                               default=[teacher_map[tid] for tid in assigned_ids if tid in teacher_map],
                                               key=f"teachers_{sub_id}")
                        if st.button("💾 حفظ", key=f"save_{sub_id}"):
                            try:
                                sel_ids = [avail[n] for n in sel]
                                sb.table("subject_teachers").delete().eq("subject_id", sub_id).execute()
                                if sel_ids:
                                    sb.table("subject_teachers").insert(
                                        [{"subject_id": sub_id, "teacher_id": tid} for tid in sel_ids]
                                    ).execute()
                                st.success("✅ تم الحفظ")
                                st.rerun()
                            except Exception as e:
                                st.error(f"خطأ: {e}")
                    else:
                        st.info("لا يوجد معلمون.")
                with c2:
                    if st.button("🗑️ حذف", key=f"ds_{sub_id}", type="secondary"):
                        sb.table("subjects").update({"is_active": False}).eq("id", sub_id).execute()
                        st.rerun()
    except Exception as e:
        st.error(f"خطأ: {e}")


# ─────────────────────────────────────────────────────────
# ACCESS CODES
# ─────────────────────────────────────────────────────────
def _access_codes(sb, lang, uid):
    st.title("🔑 أكواد الوصول")
    from app.services.access_code_service import AccessCodeService
    from app.reports.excel_exporter import export_codes_excel
    svc  = AccessCodeService()
    subs = sb.table("subjects").select("id, name_ar").eq("is_active", True).execute().data or []
    if not subs:
        st.info("لا توجد مواد.")
        return

    sub_opts = {s["name_ar"]: s["id"] for s in subs}
    chosen   = st.selectbox("اختر المادة", list(sub_opts.keys()))
    sid      = sub_opts[chosen]

    with st.expander("🔑 توليد أكواد جديدة"):
        with st.form("gen_codes"):
            c1, c2, c3 = st.columns(3)
            with c1:
                b_name  = st.text_input("اسم الباتش", value="Batch 1")
                b_desc  = st.text_input("الوصف (اختياري)")
            with c2:
                qty      = st.number_input("عدد الأكواد", 1, 500, 10)
                max_uses = st.number_input("الحد الأقصى للاستخدام", 0, 100, 1)
            with c3:
                use_exp  = st.checkbox("تحديد تاريخ انتهاء")
                exp_date = st.date_input("تاريخ الانتهاء") if use_exp else None
            if st.form_submit_button("🚀 توليد", use_container_width=True):
                with st.spinner("جاري التوليد..."):
                    try:
                        svc.create_batch(sid, b_name, b_desc, int(qty), int(max_uses), exp_date, uid)
                        st.success(f"✅ تم توليد {qty} كود")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

    batches = svc.get_subject_batches(sid)
    for batch in batches:
        stats = svc.get_batch_analytics(batch["id"])
        with st.expander(f"📦 {batch['batch_name']} | ✅ {stats.get('used',0)} | 🔵 {stats.get('unused',0)} متاح"):
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("إجمالي", stats.get("total", 0))
            c2.metric("مستخدم", stats.get("used", 0))
            c3.metric("متاح",   stats.get("unused", 0))
            c4.metric("معدل",   f"{stats.get('redemption_rate',0)}%")

            codes = svc.get_batch_codes(batch["id"])
            buf, fname = export_codes_excel(codes, batch["batch_name"], chosen)
            st.download_button("📥 Excel", buf, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"xl_{batch['id']}")
            df = pd.DataFrame([{
                "الكود":     c["code"],
                "الاستخدام": f"{c['uses_count']}/{c['max_uses'] or '∞'}",
                "نشط":      "✅" if c["is_active"] else "❌",
                "الانتهاء": str(c["expires_at"])[:10] if c["expires_at"] else "—",
            } for c in codes])
            st.dataframe(df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────
def _results(sb, lang):
    st.title("📈 النتائج")
    try:
        res = sb.table("results") \
                .select("*, profiles(full_name), exams(title)") \
                .order("submitted_at", desc=True).limit(200).execute()
        if res.data:
            df = pd.DataFrame([{
                "الطالب":   (r.get("profiles") or {}).get("full_name", "—"),
                "الامتحان": (r.get("exams") or {}).get("title", "—"),
                "الدرجة":   f"{r['score']}/{r['total']}",
                "%":        f"{round(r['score']/r['total']*100) if r['total'] else 0}%",
                "التاريخ":  str(r["submitted_at"])[:10],
            } for r in res.data])
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("لا توجد نتائج")
    except Exception as e:
        st.error(f"خطأ: {e}")


# ─────────────────────────────────────────────────────────
# AUDIT
# ─────────────────────────────────────────────────────────
def _audit(sb, lang):
    st.title("📋 سجل الأحداث")
    try:
        res = sb.table("audit_logs") \
                .select("*, profiles(full_name)") \
                .order("created_at", desc=True).limit(300).execute()
        if res.data:
            df = pd.DataFrame([{
                "المستخدم": (r.get("profiles") or {}).get("full_name", "—"),
                "الإجراء":  r["action"],
                "النوع":    r.get("entity", "—"),
                "التفاصيل": str(r.get("metadata", {}))[:80],
                "الوقت":    str(r["created_at"])[:19],
            } for r in res.data])
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("لا توجد أحداث")
    except Exception as e:
        st.error(f"خطأ: {e}")
